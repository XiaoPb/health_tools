"""check 报告 XLSX 的分类聚合与工作簿写出。

分类语义与 ``check_operation`` 的主要异常匹配保持一致，但为独立实现：必须按
状态列匹配，不能从 主要异常项 文本反推。未命中任何规则的报告行不进入任何分类
表，只保留在总表中（分类说明与 category 表只覆盖 issue_priority 命中的行）。
``write_check_workbook`` 基于这些聚合结果写出固定顺序的多工作表 XLSX 报告。
"""

import unicodedata
from collections import Counter
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Sequence, Set, Tuple, TypedDict

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]

# 稳定异常组标识到 (状态列, 期望值, 目录分类) 的映射；匹配顺序由调用方传入的
# issue_priority 决定。reference 组的两个状态列按 OR 语义匹配。
_ROW_RULES: Dict[str, Tuple[Tuple[str, ...], str, str]] = {
    "frame_fail": (("帧完整性(结果)",), "FAIL", "frame"),
    "range_fail": (("数据范围(结果)",), "FAIL", "range"),
    "acc_fail": (("ACC异常(结果)",), "FAIL", "acc_fail"),
    "timestamp_fail": (("时间戳间隔(结果)",), "FAIL", "timestamp"),
    "frame_warning": (("帧完整性(结果)",), "WARNING", "frame_warning"),
    "reference_fail": (("心率金标(结果)", "血氧金标(结果)"), "FAIL", "reference"),
    "acc_warning": (("ACC异常(结果)",), "WARNING", "acc_warning"),
    "center_fail": (("数据居中(结果)",), "FAIL", "center"),
}

# 稳定异常组标识到目录分类的映射；与 _ROW_RULES 中的 category 保持一致，
# 供 expand_issue_category_order 按优先级展开分类顺序使用。
_ISSUE_CATEGORY: Dict[str, str] = {rule_id: rule[2] for rule_id, rule in _ROW_RULES.items()}

# 目录分类到稳定异常组标识的反向映射，供分类说明表的 优先级 列使用；多个
# rule_id 映射到同一 category 时取最先出现者（first-wins，防御性；当前 8 个
# category 互不相同）。准确度标记分类不在其中，写表时统一回退为 "accuracy"。
_ISSUE_ID_BY_CATEGORY: Dict[str, str] = {}
for _rule_id, _category in _ISSUE_CATEGORY.items():
    if _category not in _ISSUE_ID_BY_CATEGORY:
        _ISSUE_ID_BY_CATEGORY[_category] = _rule_id

# Excel 工作表名非法字符（openpyxl 与 Excel 均不允许）。
_SHEET_ILLEGAL_CHARS = frozenset("[]:*?/\\")

# 分类说明表的固定列。
_CATEGORY_SUMMARY_COLUMNS: Tuple[str, ...] = (
    "优先级",
    "异常类别",
    "文件数",
    "占比",
    "判断条件",
    "场景分类占比",
    "人员占比",
    "说明信息",
)

# Excel 工作表名最大长度与保留名（Excel 不允许名为 History 的工作表）。
_SHEET_TITLE_MAX_LENGTH = 31
_SHEET_TITLE_FALLBACK = "sheet"
_SHEET_RESERVED_NAMES = frozenset({"History"})


class CategorySummary(TypedDict):
    """build_category_summary 返回的分类说明表字段。"""

    category: str
    condition: str
    explanation: str
    count: int
    ratio: float
    scene_distribution: str
    person_distribution: str


# 与 check_operation.issue_category_order 不同，本函数只返回 rows 中实际命中的
# category，且不附加 agc/ipd/total_fail/normal 兜底类别。
def _ordered_present_categories(
    present: Set[str],
    issue_priority: Sequence[str],
    accuracy_categories: Sequence[str],
) -> Tuple[str, ...]:
    """按 issue_priority 展开 present 中实际出现的分类顺序（公共排序逻辑）。

    稳定异常组映射为对应目录分类；``accuracy`` 在外层位置展开为按
    ``accuracy_categories`` 声明顺序排列的准确度分类。不在 present 中的分类不
    输出；未知标识被跳过，不产生任何 category。
    """
    categories: List[str] = []
    for rule_id in issue_priority:
        if rule_id == "accuracy":
            categories.extend(
                value
                for value in accuracy_categories
                if value in present and value not in categories
            )
            continue
        category = _ISSUE_CATEGORY.get(rule_id, rule_id)
        if category in present and category not in categories:
            categories.append(category)
    return tuple(categories)


def expand_issue_category_order(
    issue_priority: Sequence[str],
    rows: Sequence[Mapping[str, object]],
    accuracy_categories: Sequence[str] = (),
) -> Tuple[str, ...]:
    """按 issue_priority 展开为实际出现在 rows 中的分类顺序。

    稳定异常组映射为对应目录分类；``accuracy`` 在外层位置展开为实际出现在
    rows 中且按 ``accuracy_categories`` 声明顺序排列的准确度分类。未出现在
    rows 中的分类不输出；未知标识被跳过，不产生任何 category。
    """
    present = {
        category for row in rows for category in (_row_category(row, issue_priority),) if category
    }
    return _ordered_present_categories(present, issue_priority, accuracy_categories)


def build_category_summary(
    rows: Sequence[Mapping[str, object]],
    category: str,
    condition: str,
    explanation: str,
    total_count: int,
) -> CategorySummary:
    """统计单个分类命中行的聚合结果。

    返回分类说明表所需字段：count（命中行数，整数）、ratio（count/total_count
    数值）、scene_distribution（按 场景分类 计数）、person_distribution（按 姓名
    计数）、condition/explanation 原样保留，以及 category 供后续工作表使用。
    占比的分母是总表行数；场景与人员占比的分母是该分类命中行数；空姓名/场景
    统一为 default。total_count 是总表行数（count <= total_count 的预期前提）。
    """
    count = len(rows)
    scene_counter = _count_field(rows, "场景分类")
    person_counter = _count_field(rows, "姓名")
    return {
        "category": category,
        "count": count,
        "ratio": count / total_count if total_count else 0.0,
        "scene_distribution": _format_distribution(scene_counter, count),
        "person_distribution": _format_distribution(person_counter, count),
        "condition": condition,
        "explanation": explanation,
    }


def _row_category(
    row: Mapping[str, object],
    issue_priority: Sequence[str],
) -> str:
    """按状态列匹配行的目录分类；未命中返回空字符串。

    必须按状态列（如 帧完整性(结果)==FAIL）匹配，不能从 主要异常项 文本反推；
    reference 组命中任一金标列即返回 reference。``accuracy`` 在 准确度标定分类
    非空时返回该值本身。返回 "" 表示未命中。
    """
    for rule_id in issue_priority:
        if rule_id == "accuracy":
            value = str(row.get("准确度标定分类", "") or "").strip()
            if value:
                return value
            continue
        rule = _ROW_RULES.get(rule_id)
        if rule is None:
            continue
        columns, expected, category = rule
        if any(str(row.get(column, "") or "").strip().upper() == expected for column in columns):
            return category
    return ""


def _format_distribution(counter: Counter, total: int) -> str:
    """把分类计数格式化为 "name: count (pct.00%)"，条目间以 ", " 连接。

    空名统一归一为 default；total 为 0 时占比按 0.00% 处理，避免除零崩溃。
    """
    items = []
    for name, count in counter.items():
        display = name if name else "default"
        percent = count / total * 100 if total else 0.0
        items.append(f"{display}: {count} ({percent:.2f}%)")
    return ", ".join(items)


def _count_field(rows: Sequence[Mapping[str, object]], column: str) -> Counter:
    """按列统计字段出现次数，空值与纯空白值统一归一为 default。"""
    counter: Counter = Counter()
    for row in rows:
        value = str(row.get(column, "") or "").strip()
        counter[value if value else "default"] += 1
    return counter


def _safe_sheet_title(title: str, used_names: Set[str]) -> str:
    """把原始标题转换为合法的 Excel 工作表名。

    替换非法字符 ``[ ] : * ? / \\`` 与控制字符 (0x00-0x1F) 为 ``_``，去除首尾
    空白与撇号（Excel 不允许以撇号开头或结尾），截断到 31 字符；结果为空时
    回退为 "sheet"。与已用名称冲突或命中保留名 "History" 时按创建顺序追加稳定
    后缀 ``_2``、``_3``……（首个占用原名的除外），保证重复运行结果一致。调用方
    负责把返回值加入 ``used_names``。
    """
    sanitized = (
        "".join("_" if ch in _SHEET_ILLEGAL_CHARS or ord(ch) < 0x20 else ch for ch in title)
        .strip()
        .strip("'")
    )
    base = sanitized[:_SHEET_TITLE_MAX_LENGTH] if sanitized else _SHEET_TITLE_FALLBACK
    candidate = base
    counter = 1
    while candidate in _SHEET_RESERVED_NAMES or candidate in used_names:
        counter += 1
        suffix = f"_{counter}"
        candidate = base[: _SHEET_TITLE_MAX_LENGTH - len(suffix)] + suffix
    return candidate


def _column_union(rows: Sequence[Mapping[str, object]]) -> List[str]:
    """按首见顺序求所有行键的并集，作为工作表列头。"""
    columns: List[str] = []
    seen: Set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def _display_width(text: str) -> int:
    """按 Excel 显示宽度计算：东亚全角/宽字符（F/W）按 2 列计。"""
    return sum(2 if unicodedata.east_asian_width(ch) in {"F", "W"} else 1 for ch in text)


def _write_data_sheet(
    wb: Workbook,
    title: str,
    columns: Sequence[str],
    data_rows: Sequence[Sequence[object]],
    number_formats: Optional[Mapping[int, str]] = None,
    table_index: int = 0,
) -> None:
    """创建数据工作表并写入列头与数据行，统一设置冻结、表格样式、列宽与居中。

    全部单元格水平居中、垂直居中并启用换行；数据区套用蓝色表格样式
    ``TableStyleMedium2``（自带筛选按钮）。列宽采用启发式：列头与数据前 20 行
    及后 20 行内容的 ``min(max(显示宽度) + 2, 60)``，显示宽度按东亚宽字符计 2 列。
    ``number_formats`` 为 {1 起始列序号: 数字格式}，仅对数据行生效。
    """
    ws = wb.create_sheet(title=title)
    if columns:
        ws.append(list(columns))
        for row in data_rows:
            ws.append(list(row))
    ws.freeze_panes = "A2"
    if columns:
        last_column = get_column_letter(len(columns))
        table = Table(
            displayName=f"CheckTable{table_index}",
            ref=f"A1:{last_column}{ws.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        for index, column in enumerate(columns, start=1):
            sample = [column]
            head = list(data_rows[:20])
            tail = list(data_rows[-20:]) if len(data_rows) > 40 else []
            for row in head + tail:
                sample.append(str(row[index - 1]) if index - 1 < len(row) else "")
            ws.column_dimensions[get_column_letter(index)].width = min(
                max(_display_width(value) for value in sample) + 2, 60
            )
    if number_formats:
        for row_index in range(2, ws.max_row + 1):
            for column_index, number_format in number_formats.items():
                ws.cell(row=row_index, column=column_index).number_format = number_format
    for row_index in range(1, ws.max_row + 1):
        for column_index in range(1, ws.max_column + 1):
            ws.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )


def write_check_workbook(
    output: Path,
    rows: Sequence[Mapping[str, object]],
    compact_rows: Sequence[Mapping[str, object]],
    issue_priority: Sequence[str],
    accuracy_categories: Sequence[str] = (),
    category_descriptions: Optional[Mapping[str, Tuple[str, str]]] = None,
) -> None:
    """写出固定顺序的多工作表 check XLSX 报告。

    固定顺序：总表 → 分类说明 → （每个实际命中分类一张表，按
    ``expand_issue_category_order`` 顺序）→ 精简总表。输出目录不存在时自动创建。

    - 总表：全部行与全部列（首见顺序并集，缺键补空串），首行为列头。
    - 分类说明：每类一行统计，列为 ``_CATEGORY_SUMMARY_COLUMNS``。
    - 分类表：仅含 ``_row_category`` 命中该分类的行，列与总表一致。
    - 精简总表：原样写出 ``compact_rows`` 的列头与内容。

    ``category_descriptions`` 的键是**展开后的分类名**（如 "frame"，或准确度标记
    分类值如 "accuracy_online_low"），不是规则标识（如 "frame_fail"）；缺失的键
    回退为 ("", "")。

    空输入仍会生成 总表/分类说明/精简总表，分类表只覆盖实际命中的分类。
    """
    output.parent.mkdir(parents=True, exist_ok=True)
    category_descriptions = category_descriptions or {}
    total_columns = _column_union(rows)
    # 单趟分类：每行只计算一次 _row_category，写入其唯一所属分类桶；同趟收集
    # 实际出现的分类集合，供分类顺序展开过滤使用（总体 O(rows × priority)）。
    rows_by_category: Dict[str, List[Mapping[str, object]]] = {}
    present: Set[str] = set()
    for row in rows:
        category = _row_category(row, issue_priority)
        if category:
            present.add(category)
            rows_by_category.setdefault(category, []).append(row)
    categories = _ordered_present_categories(present, issue_priority, accuracy_categories)

    wb = Workbook()
    wb.remove(wb.active)  # 移除默认空表，按固定顺序重建
    # 预占三个固定工作表名，避免分类表与它们重名（openpyxl 会自动改名破坏固定名）。
    used_names: Set[str] = {"总表", "分类说明", "精简总表"}
    table_index = 0

    # 总表
    _write_data_sheet(
        wb,
        "总表",
        total_columns,
        [[row.get(column, "") for column in total_columns] for row in rows],
        table_index=table_index,
    )
    table_index += 1

    # 分类说明
    summary_rows: List[List[object]] = []
    for category in categories:
        condition, explanation = category_descriptions.get(category, ("", ""))
        summary = build_category_summary(
            rows_by_category[category],
            category,
            condition,
            explanation,
            total_count=len(rows),
        )
        summary_rows.append(
            [
                _ISSUE_ID_BY_CATEGORY.get(category, "accuracy"),
                category,
                summary["count"],
                summary["ratio"],
                condition,
                summary["scene_distribution"],
                summary["person_distribution"],
                explanation,
            ]
        )
    _write_data_sheet(
        wb,
        "分类说明",
        _CATEGORY_SUMMARY_COLUMNS,
        summary_rows,
        number_formats={3: "0", 4: "0.00%"},
        table_index=table_index,
    )
    table_index += 1

    # 分类表（每个实际命中的 category 一张，标题经 _safe_sheet_title 清洗）
    for category in categories:
        title = _safe_sheet_title(category, used_names)
        used_names.add(title)
        _write_data_sheet(
            wb,
            title,
            total_columns,
            [
                [row.get(column, "") for column in total_columns]
                for row in rows_by_category[category]
            ],
            table_index=table_index,
        )
        table_index += 1

    # 精简总表（始终最后）
    compact_columns = _column_union(compact_rows)
    _write_data_sheet(
        wb,
        "精简总表",
        compact_columns,
        [[row.get(column, "") for column in compact_columns] for row in compact_rows],
        table_index=table_index,
    )

    wb.save(output)
