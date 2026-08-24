"""check XLSX 报告分类聚合与工作簿写出的单元测试。"""

import csv
from collections import Counter

import pytest
from openpyxl import load_workbook

from health_tools.api.check_operation import COMPACT_HEADER, _report_rows_for_xlsx, run_check
from health_tools.api.models import CheckRequest
from health_tools.core.checker import FileCheckReport
from health_tools.models.rules import AccuracyMarkRule
from health_tools.utils.check_xlsx import (
    _count_field,
    _format_distribution,
    _row_category,
    _safe_sheet_title,
    build_category_summary,
    expand_issue_category_order,
    write_check_workbook,
)


def test_expand_issue_category_order_expands_accuracy_marks_in_place():
    rows = [{"准确度标定分类": "accuracy_online_low"}, {"准确度标定分类": "accuracy_comp_low"}]
    assert expand_issue_category_order(
        ("accuracy", "frame_fail"), rows, ("accuracy_online_low", "accuracy_comp_low")
    ) == ("accuracy_online_low", "accuracy_comp_low")


def test_build_category_summary_calculates_ratios():
    rows = [
        {"场景分类": "rest", "姓名": "alice"},
        {"场景分类": "rest", "姓名": "bob"},
    ]
    summary = build_category_summary(
        rows,
        category="frame",
        condition="帧完整性(结果)=FAIL",
        explanation="帧不完整",
        total_count=2,
    )
    assert summary["count"] == 2
    assert summary["ratio"] == 1.0
    assert summary["scene_distribution"] == "rest: 2 (100.00%)"
    assert "alice: 1 (50.00%)" in summary["person_distribution"]


def test_expand_issue_category_order_keeps_stable_category_mapping():
    rows = [
        {"帧完整性(结果)": "FAIL"},
        {"数据范围(结果)": "FAIL"},
        {"ACC异常(结果)": "FAIL"},
        {"时间戳间隔(结果)": "FAIL"},
        {"帧完整性(结果)": "WARNING"},
        {"心率金标(结果)": "FAIL"},
        {"ACC异常(结果)": "WARNING"},
        {"数据居中(结果)": "FAIL"},
    ]
    assert expand_issue_category_order(
        (
            "frame_fail",
            "range_fail",
            "acc_fail",
            "timestamp_fail",
            "frame_warning",
            "reference_fail",
            "acc_warning",
            "center_fail",
        ),
        rows,
        (),
    ) == (
        "frame",
        "range",
        "acc_fail",
        "timestamp",
        "frame_warning",
        "reference",
        "acc_warning",
        "center",
    )


def test_expand_issue_category_order_filters_categories_missing_from_rows():
    rows = [{"帧完整性(结果)": "FAIL"}]
    assert expand_issue_category_order(
        ("frame_fail", "range_fail", "timestamp_fail"), rows, ()
    ) == ("frame",)


def test_expand_issue_category_order_uses_accuracy_declaration_order():
    rows = [
        {"准确度标定分类": "accuracy_comp_low"},
        {"准确度标定分类": "accuracy_online_low"},
    ]
    assert expand_issue_category_order(
        ("accuracy",), rows, ("accuracy_online_low", "accuracy_comp_low")
    ) == ("accuracy_online_low", "accuracy_comp_low")


def test_row_category_matches_status_columns_in_priority_order():
    row = {
        "帧完整性(结果)": "FAIL",
        "数据范围(结果)": "FAIL",
        "ACC异常(结果)": "FAIL",
        "总异常(结果)": "FAIL",
    }
    assert _row_category(row, ("frame_fail", "range_fail", "acc_fail")) == "frame"
    assert _row_category(row, ("range_fail", "frame_fail")) == "range"
    assert _row_category(row, ("acc_fail",)) == "acc_fail"


def test_row_category_reference_matches_either_reference_column():
    for column in ("心率金标(结果)", "血氧金标(结果)"):
        row = {"总异常(结果)": "FAIL", column: "FAIL", "数据居中(结果)": "FAIL"}
        assert _row_category(row, ("reference_fail", "center_fail")) == "reference"


def test_row_category_accuracy_returns_mark_category_value():
    row = {"准确度标定分类": "accuracy_online_low", "帧完整性(结果)": "FAIL"}
    assert _row_category(row, ("accuracy", "frame_fail")) == "accuracy_online_low"
    assert _row_category(row, ("frame_fail", "accuracy")) == "frame"


def test_row_category_ignores_primary_issue_text():
    row = {"主要异常项": "帧不完整", "帧完整性(结果)": "PASS"}
    assert _row_category(row, ("frame_fail",)) == ""


def test_format_distribution_normalizes_empty_name_and_zero_total():
    counter = Counter({"": 2, "rest": 1})
    assert _format_distribution(counter, 3) == "default: 2 (66.67%), rest: 1 (33.33%)"
    assert _format_distribution(counter, 0) == "default: 2 (0.00%), rest: 1 (0.00%)"


def test_expand_issue_category_order_dedupes_duplicate_issue_ids():
    rows = [{"帧完整性(结果)": "FAIL"}, {"准确度标定分类": "accuracy_online_low"}]
    assert expand_issue_category_order(
        ("frame_fail", "frame_fail", "accuracy"), rows, ("accuracy_online_low",)
    ) == ("frame", "accuracy_online_low")


def test_expand_issue_category_order_dedupes_duplicate_accuracy_categories():
    rows = [{"准确度标定分类": "accuracy_online_low"}]
    assert expand_issue_category_order(
        ("accuracy",), rows, ("accuracy_online_low", "accuracy_online_low")
    ) == ("accuracy_online_low",)


def test_expand_issue_category_order_empty_rows_returns_empty_tuple():
    assert (
        expand_issue_category_order(("accuracy", "frame_fail"), [], ("accuracy_online_low",)) == ()
    )


def test_build_category_summary_zero_total_count_ratio_is_zero():
    rows = [{"场景分类": "rest", "姓名": "alice"}]
    summary = build_category_summary(rows, "frame", "cond", "expl", total_count=0)
    assert summary["count"] == 1
    assert summary["ratio"] == 0.0
    assert summary["scene_distribution"] == "rest: 1 (100.00%)"


def test_build_category_summary_empty_rows_zero_counts():
    summary = build_category_summary([], "frame", "cond", "expl", total_count=5)
    assert summary["count"] == 0
    assert summary["ratio"] == 0.0
    assert summary["scene_distribution"] == ""
    assert summary["person_distribution"] == ""


def test_row_category_missing_status_columns_returns_empty():
    row = {"姓名": "alice", "主要异常项": "帧不完整"}
    assert _row_category(row, ("frame_fail", "range_fail")) == ""


def test_row_category_matches_status_values_case_insensitively():
    row = {"帧完整性(结果)": "fail", "数据范围(结果)": "FAIL"}
    assert _row_category(row, ("frame_fail", "range_fail")) == "frame"
    assert _row_category(row, ("range_fail", "frame_fail")) == "range"


def test_row_category_skips_unknown_rule_ids():
    row = {"帧完整性(结果)": "FAIL"}
    assert _row_category(row, ("unknown_rule", "frame_fail")) == "frame"
    assert _row_category(row, ("unknown_rule", "other_unknown")) == ""


def test_count_field_normalizes_whitespace_only_values_to_default():
    rows = [
        {"场景分类": "   ", "姓名": "  "},
        {"场景分类": "rest", "姓名": "alice"},
    ]
    assert _count_field(rows, "场景分类") == Counter({"default": 1, "rest": 1})
    assert _count_field(rows, "姓名") == Counter({"default": 1, "alice": 1})


def test_build_category_summary_normalizes_whitespace_only_distribution_values():
    rows = [{"场景分类": " \t ", "姓名": "  "}]
    summary = build_category_summary(rows, "frame", "cond", "expl", total_count=1)
    assert summary["scene_distribution"] == "default: 1 (100.00%)"
    assert summary["person_distribution"] == "default: 1 (100.00%)"


def test_write_check_workbook_creates_ordered_sheets(tmp_path):
    rows = [
        {
            "文件名": "a.csv",
            "总异常(结果)": "FAIL",
            "帧完整性(结果)": "FAIL",
            "场景分类": "rest",
            "姓名": "alice",
            "准确度标定分类": "",
        }
    ]
    output = tmp_path / "check_report.xlsx"
    write_check_workbook(
        output,
        rows,
        [{"文件名": "a.csv", "检查项": "帧完整性", "状态": "FAIL"}],
        issue_priority=("frame_fail", "accuracy"),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "frame", "精简总表"]
    assert book["总表"]["A2"].value == "a.csv"
    assert book["frame"]["A2"].value == "a.csv"


def test_write_check_workbook_keeps_base_sheets_when_empty(tmp_path):
    output = tmp_path / "empty.xlsx"
    write_check_workbook(
        output,
        [],
        [],
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "精简总表"]
    assert book["分类说明"]["A1"].value == "优先级"
    assert book["总表"].max_row == 1


def test_safe_sheet_title_sanitizes_illegal_characters_and_truncates():
    assert _safe_sheet_title("a[b]:c*d?e/f\\g", set()) == "a_b__c_d_e_f_g"
    assert _safe_sheet_title("  frame  ", set()) == "frame"
    assert _safe_sheet_title("x" * 40, set()) == "x" * 31
    assert _safe_sheet_title("", set()) == "sheet"
    assert _safe_sheet_title("   ", set()) == "sheet"


def test_safe_sheet_title_appends_suffix_on_conflict_and_reserved_history():
    assert _safe_sheet_title("frame", {"frame"}) == "frame_2"
    assert _safe_sheet_title("frame", {"frame", "frame_2"}) == "frame_3"
    assert _safe_sheet_title("History", set()) == "History_2"
    assert _safe_sheet_title("History", {"History_2"}) == "History_3"


def test_safe_sheet_title_strips_control_chars_and_apostrophes():
    assert _safe_sheet_title("a\x00b\x1f", set()) == "a_b_"
    assert _safe_sheet_title("line\nbreak", set()) == "line_break"
    assert _safe_sheet_title("'frame'", set()) == "frame"
    assert _safe_sheet_title("  'frame'  ", set()) == "frame"
    assert _safe_sheet_title("'", set()) == "sheet"


def test_safe_sheet_title_truncated_base_with_suffix_stays_within_limit():
    base = "x" * 31
    assert _safe_sheet_title(base, {base}) == "x" * 29 + "_2"
    assert len(_safe_sheet_title(base, {base})) == 31


def test_write_check_workbook_applies_numeric_formats(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL", "场景分类": "rest", "姓名": "alice"},
        {"文件名": "b.csv", "帧完整性(结果)": "FAIL", "场景分类": "walk", "姓名": "bob"},
        {"文件名": "c.csv", "帧完整性(结果)": "PASS", "场景分类": "rest", "姓名": "carol"},
    ]
    output = tmp_path / "formats.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    summary = book["分类说明"]
    assert summary["A2"].value == "frame_fail"
    assert summary["B2"].value == "frame"
    assert summary["C2"].value == 2
    assert summary["C2"].number_format == "0"
    assert summary["D2"].value == 2 / 3
    assert summary["D2"].number_format == "0.00%"
    assert summary["E2"].value == "帧完整性(结果)=FAIL"
    assert summary["H2"].value == "帧不完整"


def test_write_check_workbook_sets_freeze_panes_and_auto_filter(tmp_path):
    rows = [{"文件名": "a.csv", "帧完整性(结果)": "FAIL"}]
    output = tmp_path / "frozen.xlsx"
    write_check_workbook(
        output,
        rows,
        [{"文件名": "a.csv"}],
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    for name in ("总表", "分类说明", "frame", "精简总表"):
        ws = book[name]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None


def test_write_check_workbook_creates_category_sheets_only_for_hit_categories(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL"},
        {"文件名": "b.csv", "数据范围(结果)": "FAIL"},
        {"文件名": "c.csv", "帧完整性(结果)": "PASS"},
    ]
    output = tmp_path / "categories.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail", "range_fail", "timestamp_fail"),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "frame", "range", "精简总表"]
    assert book["frame"].max_row == 2
    assert book["range"]["A2"].value == "b.csv"


def test_write_check_workbook_accuracy_sheet_sanitizes_title_keeps_data(tmp_path):
    rows = [{"文件名": "a.csv", "准确度标定分类": "acc[online]:low", "姓名": "alice"}]
    output = tmp_path / "accuracy.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("accuracy",),
        accuracy_categories=("acc[online]:low",),
        category_descriptions={},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "acc_online__low", "精简总表"]
    ws = book["acc_online__low"]
    headers = [cell.value for cell in ws[1]]
    mark_column = headers.index("准确度标定分类") + 1
    assert ws.cell(row=2, column=mark_column).value == "acc[online]:low"
    assert book["分类说明"]["A2"].value == "accuracy"
    assert book["分类说明"]["B2"].value == "acc[online]:low"


def test_write_check_workbook_compact_sheet_matches_compact_rows(tmp_path):
    compact_rows = [
        {"文件名": "a.csv", "检查项": "帧完整性", "状态": "FAIL"},
        {"文件名": "b.csv", "检查项": "数据范围", "状态": "FAIL"},
    ]
    output = tmp_path / "compact.xlsx"
    write_check_workbook(
        output,
        [],
        compact_rows,
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={},
    )
    book = load_workbook(output)
    ws = book["精简总表"]
    assert [cell.value for cell in ws[1]] == ["文件名", "检查项", "状态"]
    assert [cell.value for cell in ws[2]] == ["a.csv", "帧完整性", "FAIL"]
    assert [cell.value for cell in ws[3]] == ["b.csv", "数据范围", "FAIL"]
    assert ws.max_row == 3


def test_write_check_workbook_category_matching_fixed_names_gets_suffix(tmp_path):
    rows = [{"文件名": "a.csv", "准确度标定分类": "精简总表"}]
    output = tmp_path / "reserved.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("accuracy",),
        accuracy_categories=("精简总表",),
        category_descriptions={},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "精简总表_2", "精简总表"]


def test_write_check_workbook_heterogeneous_keys_column_union_first_seen(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL"},
        {"数据范围(结果)": "FAIL", "姓名": "bob"},
        {"帧完整性(结果)": "PASS"},
    ]
    output = tmp_path / "union.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail", "range_fail"),
        accuracy_categories=(),
        category_descriptions={},
    )
    book = load_workbook(output)
    ws = book["总表"]
    assert [cell.value for cell in ws[1]] == ["文件名", "帧完整性(结果)", "数据范围(结果)", "姓名"]
    # 缺键按列序补空串写出；openpyxl 把空串单元格写为空单元格，读回为 None
    assert [cell.value for cell in ws[2]] == ["a.csv", "FAIL", None, None]
    assert [cell.value for cell in ws[3]] == [None, None, "FAIL", "bob"]
    assert [cell.value for cell in ws[4]] == [None, "PASS", None, None]


def test_write_check_workbook_non_string_values_round_trip(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL", "数量": 3, "占比": 0.75, "备注": None},
        {"文件名": "b.csv", "帧完整性(结果)": "FAIL", "数量": 1, "占比": 0.25, "备注": None},
    ]
    output = tmp_path / "values.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    headers = [cell.value for cell in book["总表"][1]]
    assert headers == ["文件名", "帧完整性(结果)", "数量", "占比", "备注"]
    assert [cell.value for cell in book["总表"][2]] == ["a.csv", "FAIL", 3, 0.75, None]
    assert [cell.value for cell in book["总表"][3]] == ["b.csv", "FAIL", 1, 0.25, None]
    assert [cell.value for cell in book["frame"][2]] == ["a.csv", "FAIL", 3, 0.75, None]
    assert [cell.value for cell in book["frame"][3]] == ["b.csv", "FAIL", 1, 0.25, None]


def test_write_check_workbook_deterministic_output(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL", "场景分类": "rest", "姓名": "alice"},
        {"文件名": "b.csv", "数据范围(结果)": "FAIL", "场景分类": "walk", "姓名": "bob"},
    ]
    compact_rows = [{"文件名": "a.csv", "检查项": "帧完整性", "状态": "FAIL"}]
    out1 = tmp_path / "deterministic_1.xlsx"
    out2 = tmp_path / "deterministic_2.xlsx"
    write_check_workbook(
        out1,
        rows,
        compact_rows,
        issue_priority=("frame_fail", "range_fail"),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    write_check_workbook(
        out2,
        rows,
        compact_rows,
        issue_priority=("frame_fail", "range_fail"),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book1 = load_workbook(out1)
    book2 = load_workbook(out2)
    assert book1.sheetnames == book2.sheetnames
    for name in book1.sheetnames:
        ws1, ws2 = book1[name], book2[name]
        assert ws1.max_row == ws2.max_row
        assert ws1.max_column == ws2.max_column
        for row in ws1.iter_rows():
            for cell in row:
                assert cell.value == ws2[cell.coordinate].value


def test_write_check_workbook_sanitized_title_collision_gets_stable_suffix(tmp_path):
    rows = [
        {"文件名": "a.csv", "准确度标定分类": "a/b"},
        {"文件名": "b.csv", "准确度标定分类": "a_b"},
    ]
    output = tmp_path / "collision.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("accuracy",),
        accuracy_categories=("a/b", "a_b"),
        category_descriptions={},
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "a_b", "a_b_2", "精简总表"]
    for name, mark in (("a_b", "a/b"), ("a_b_2", "a_b")):
        ws = book[name]
        headers = [cell.value for cell in ws[1]]
        column = headers.index("准确度标定分类") + 1
        assert ws.cell(row=2, column=column).value == mark


def test_write_check_workbook_auto_filter_ref_exact(tmp_path):
    rows = [
        {
            "文件名": "a.csv",
            "帧完整性(结果)": "FAIL",
            "场景分类": "rest",
            "姓名": "alice",
            "总异常(结果)": "FAIL",
        },
        {
            "文件名": "b.csv",
            "帧完整性(结果)": "FAIL",
            "场景分类": "walk",
            "姓名": "bob",
            "总异常(结果)": "FAIL",
        },
    ]
    output = tmp_path / "filter.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail",),
        accuracy_categories=(),
        category_descriptions={"frame": ("帧完整性(结果)=FAIL", "帧不完整")},
    )
    book = load_workbook(output)
    ws = book["总表"]
    assert ws.auto_filter.ref == f"A1:E{ws.max_row}"
    assert ws.auto_filter.ref == "A1:E3"


def test_write_check_workbook_multi_category_summary_denominators(tmp_path):
    rows = [
        {"文件名": "a.csv", "帧完整性(结果)": "FAIL", "场景分类": "rest", "姓名": "alice"},
        {"文件名": "b.csv", "数据范围(结果)": "FAIL", "场景分类": "walk", "姓名": "bob"},
        {"文件名": "c.csv", "帧完整性(结果)": "FAIL", "场景分类": "walk", "姓名": "carol"},
    ]
    output = tmp_path / "multi.xlsx"
    write_check_workbook(
        output,
        rows,
        [],
        issue_priority=("frame_fail", "range_fail"),
        accuracy_categories=(),
        category_descriptions={
            "frame": ("帧完整性(结果)=FAIL", "帧不完整"),
            "range": ("数据范围(结果)=FAIL", "数据越界"),
        },
    )
    book = load_workbook(output)
    summary = book["分类说明"]
    # frame 命中 a/c 两行：占比分母为全部行数 3，场景/人员占比分母为自身命中数 2
    assert summary["B2"].value == "frame"
    assert summary["C2"].value == 2
    assert summary["D2"].value == 2 / 3
    assert summary["F2"].value == "rest: 1 (50.00%), walk: 1 (50.00%)"
    assert summary["G2"].value == "alice: 1 (50.00%), carol: 1 (50.00%)"
    # range 命中 b 一行
    assert summary["B3"].value == "range"
    assert summary["C3"].value == 1
    assert summary["D3"].value == 1 / 3
    assert summary["F3"].value == "walk: 1 (100.00%)"
    assert summary["G3"].value == "bob: 1 (100.00%)"
    # 每行只属于一个分类：文件数之和等于总行数
    assert summary["C2"].value + summary["C3"].value == len(rows)


def write_check_fixture(tmp_path, columns, rows, name="check_fixture.csv"):
    """按 gh3220 CSV 布局写出 fixture：info 行 1、表头行 2、数据自第 3 行起。"""
    path = tmp_path / name
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Version: GH3220"])
        writer.writerow(columns)
        writer.writerows(rows)
    return path


def _accuracy_mark() -> AccuracyMarkRule:
    """命中 accuracy_online_low 的在线准确度标记（ref 80 vs online 50）。"""
    return AccuracyMarkRule(
        id="online_low",
        left="online.within_5",
        operator="lt",
        threshold=80.0,
        category="accuracy_online_low",
        label="Online准确度低",
    )


def _accuracy_fixture(tmp_path, name="check_fixture.csv"):
    """带准确度金标/在线列的 gh3220 fixture，命中 ``_accuracy_mark``。"""
    return write_check_fixture(
        tmp_path,
        columns=["TimeStamp", "FRAME_ID", "CH0", "REF_RESULT0", "ALGO_RESULT0"],
        rows=[
            [0, 0, 12000000, 80, 50],
            [40, 1, 12000000, 80, 50],
            [80, 2, 12000000, 80, 50],
        ],
        name=name,
    )


@pytest.fixture
def fixture_csv(tmp_path):
    """gh3220 最小可用检查数据（info 行 + 表头 + 两行数据）。"""
    return write_check_fixture(
        tmp_path,
        columns=["TimeStamp", "FRAME_ID", "ACCX", "ACCY", "ACCZ", "CH0", "CH1"],
        rows=[
            [0, 0, 100, 100, 100, 1000, 1000],
            [40, 1, 200, 200, 200, 2000, 2000],
        ],
    )


def test_xlsx_output_does_not_call_csv_writers(monkeypatch, tmp_path, fixture_csv):
    calls = {}
    monkeypatch.setattr(
        "health_tools.utils.check_xlsx.write_check_workbook",
        lambda output, rows, compact_rows, **kwargs: calls.update(output=output),
    )
    monkeypatch.setattr(
        "health_tools.api.check_operation._save_report",
        lambda *args, **kwargs: pytest.fail("CSV report writer called"),
    )
    monkeypatch.setattr(
        "health_tools.api.check_operation._save_compact_report",
        lambda *args, **kwargs: pytest.fail("CSV compact writer called"),
    )
    request = CheckRequest(
        input_path=fixture_csv,
        chip_name="gh3220",
        output_path=tmp_path / "check_report.xlsx",
    )
    run_check(request)
    assert calls["output"] == tmp_path / "check_report.xlsx"


def test_xlsx_contains_accuracy_sheet_and_final_compact_sheet(tmp_path):
    output = tmp_path / "accuracy.xlsx"
    mark = _accuracy_mark()
    accuracy_fixture_csv = _accuracy_fixture(tmp_path)
    request = CheckRequest(
        input_path=accuracy_fixture_csv,
        chip_name="gh3220",
        output_path=output,
        accuracy_enabled=True,
        accuracy_ref_column="REF_RESULT0",
        accuracy_online_column="ALGO_RESULT0",
        accuracy_marks=(mark,),
    )
    run_check(request)
    book = load_workbook(output, read_only=True)
    assert "accuracy_online_low" in book.sheetnames
    assert book.sheetnames[-1] == "精简总表"
    assert not output.with_suffix(".csv").exists()
    assert not output.with_name("accuracy_compact.csv").exists()


def test_xlsx_compact_sheet_keeps_full_columns_for_accuracy_only_batch(tmp_path):
    """仅含准确度标定命中行的批次，精简总表仍保持 COMPACT_HEADER 全列。"""
    output = tmp_path / "compact_columns.xlsx"
    request = CheckRequest(
        input_path=_accuracy_fixture(tmp_path),
        chip_name="gh3220",
        output_path=output,
        accuracy_enabled=True,
        accuracy_ref_column="REF_RESULT0",
        accuracy_online_column="ALGO_RESULT0",
        accuracy_marks=(_accuracy_mark(),),
    )
    run_check(request)
    book = load_workbook(output)
    assert book["精简总表"].max_column == len(COMPACT_HEADER)
    assert len(COMPACT_HEADER) == 32


def test_xlsx_acc_axis_output_includes_single_axis_columns(tmp_path, fixture_csv):
    """acc_axis=True 时总表（XLSX）与报告 CSV 均含单轴 ACC 列。"""
    xlsx_output = tmp_path / "axis.xlsx"
    run_check(
        CheckRequest(
            input_path=fixture_csv, chip_name="gh3220", output_path=xlsx_output, acc_axis=True
        )
    )
    book = load_workbook(xlsx_output)
    headers = [cell.value for cell in book["总表"][1]]
    for column in (
        "ACC静止X次数",
        "ACC静止X最长帧",
        "ACC静止X前10帧",
        "ACC循环Z次数",
        "ACC循环Z前10帧",
    ):
        assert column in headers
    csv_output = tmp_path / "axis.csv"
    run_check(
        CheckRequest(
            input_path=fixture_csv, chip_name="gh3220", output_path=csv_output, acc_axis=True
        )
    )
    with csv_output.open(newline="", encoding="utf-8-sig") as handle:
        csv_header = next(csv.reader(handle))
    assert "ACC静止X次数" in csv_header
    assert "ACC循环Z前10帧" in csv_header


def test_xlsx_acc_axis_acc_less_input_keeps_axis_columns(tmp_path):
    """无 ACC 列的输入开启 acc_axis=True 不崩溃，总表仍保留单轴 ACC 列。"""
    fixture = write_check_fixture(
        tmp_path,
        columns=["TimeStamp", "FRAME_ID", "CH0", "CH1"],
        rows=[
            [0, 0, 1000, 1000],
            [40, 1, 2000, 2000],
        ],
    )
    output = tmp_path / "axis_less.xlsx"
    run_check(
        CheckRequest(input_path=fixture, chip_name="gh3220", output_path=output, acc_axis=True)
    )
    book = load_workbook(output)
    headers = [cell.value for cell in book["总表"][1]]
    assert "ACC静止X次数" in headers
    assert "ACC循环Z前10帧" in headers


def test_report_rows_for_xlsx_pads_axis_columns_when_acc_missing(tmp_path):
    """acc_reports 非空但缺少当前报告的 ACC 数据时，27 列 ACC 列全部补 "-"。"""
    report = FileCheckReport(file_path=tmp_path / "a.csv", chip="gh3220")
    rows = _report_rows_for_xlsx([report], {"other.csv": object()}, tmp_path, True)
    acc_columns = [
        "ACC全零次数",
        "ACC全零最长帧",
        "ACC全零前10帧",
        "ACC静止XYZ次数",
        "ACC静止XYZ最长帧",
        "ACC静止XYZ前10帧",
        "ACC循环XYZ次数",
        "ACC循环XYZ最长帧",
        "ACC循环XYZ前10帧",
    ]
    acc_columns.extend(
        f"ACC{kind}{axis}{suffix}"
        for kind in ("静止", "循环")
        for axis in "XYZ"
        for suffix in ("次数", "最长帧", "前10帧")
    )
    assert len(acc_columns) == 27
    assert all(rows[0][column] == "-" for column in acc_columns)


def test_xlsx_uppercase_suffix_takes_xlsx_branch(monkeypatch, tmp_path, fixture_csv):
    calls = {}
    monkeypatch.setattr(
        "health_tools.utils.check_xlsx.write_check_workbook",
        lambda output, rows, compact_rows, **kwargs: calls.update(output=output),
    )
    monkeypatch.setattr(
        "health_tools.api.check_operation._save_report",
        lambda *args, **kwargs: pytest.fail("CSV report writer called"),
    )
    monkeypatch.setattr(
        "health_tools.api.check_operation._save_compact_report",
        lambda *args, **kwargs: pytest.fail("CSV compact writer called"),
    )
    request = CheckRequest(
        input_path=fixture_csv,
        chip_name="gh3220",
        output_path=tmp_path / "check_report.XLSX",
    )
    run_check(request)
    assert calls["output"] == tmp_path / "check_report.XLSX"
    assert not (tmp_path / "check_report.csv").exists()


def test_xlsx_items_only_input_writes_skip_row_without_csv_sidecars(tmp_path):
    fixture = write_check_fixture(
        tmp_path,
        columns=["Foo", "Bar"],
        rows=[
            [1, 2],
            [3, 4],
        ],
    )
    output = tmp_path / "skip.xlsx"
    run_check(CheckRequest(input_path=fixture, chip_name="gh3220", output_path=output))
    book = load_workbook(output)
    ws = book["总表"]
    assert [cell.value for cell in ws[1]] == [
        "文件名",
        "总异常(结果)",
        "主要异常项",
        "文件相对路径",
    ]
    assert ws["A2"].value == "check_fixture.csv"
    assert ws["B2"].value == "SKIP"
    assert "跳过" in ws["C2"].value
    assert not (tmp_path / "skip.csv").exists()
    assert not (tmp_path / "skip_compact.csv").exists()


def test_xlsx_accuracy_category_sheet_precedes_frame_when_both_hit(tmp_path):
    _accuracy_fixture(tmp_path, name="acc_file.csv")
    write_check_fixture(
        tmp_path,
        columns=["TimeStamp", "FRAME_ID", "CH0", "REF_RESULT0", "ALGO_RESULT0"],
        rows=[
            [0, 0, 1000, 80, 80],
            [80, 2, 2000, 80, 80],
        ],
        name="frame_file.csv",
    )
    output = tmp_path / "ordered.xlsx"
    run_check(
        CheckRequest(
            input_path=tmp_path,
            chip_name="gh3220",
            output_path=output,
            accuracy_enabled=True,
            accuracy_ref_column="REF_RESULT0",
            accuracy_online_column="ALGO_RESULT0",
            accuracy_marks=(_accuracy_mark(),),
            issue_priority=("accuracy", "frame_fail"),
        )
    )
    book = load_workbook(output)
    assert book.sheetnames == ["总表", "分类说明", "accuracy_online_low", "frame", "精简总表"]


def test_xlsx_summary_header_matches_csv_header(tmp_path, fixture_csv):
    """同一输入分别输出 .csv 与 .xlsx，总表列头必须与 CSV 列头逐列一致。"""
    csv_output = tmp_path / "report.csv"
    xlsx_output = tmp_path / "report.xlsx"
    run_check(CheckRequest(input_path=fixture_csv, chip_name="gh3220", output_path=csv_output))
    run_check(CheckRequest(input_path=fixture_csv, chip_name="gh3220", output_path=xlsx_output))
    with csv_output.open(newline="", encoding="utf-8-sig") as handle:
        csv_header = next(csv.reader(handle))
    book = load_workbook(xlsx_output)
    xlsx_header = [cell.value for cell in book["总表"][1]]
    assert xlsx_header == csv_header
